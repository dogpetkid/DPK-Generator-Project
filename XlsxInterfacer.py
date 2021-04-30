"""
This is a tool created by DPK

This tool can interface xlsx files and give out consistant datatypes.
"""

import typing

import numpy
import openpyxl
import pandas
import xlrd

# typing:   used to give types to function parameters
# pandas:   used to read the excel data
# openpyxl: used to keep the pre-existing formatting of a sheet
# xlrd:     used to catch and throw excel errors when initially reading the sheets
# numpy:    used to manipulate the inconsistant numpy data read by pandas

def ctn(col:str):
    """"Column To Number", converts a column to its index, e.g. A = 0, Z = 25, AA = 26, AAA = 702"""
    col = col.upper()
    index = 0
    i = 0
    first = True
    for letter in col[::-1]:
        index+= 26**i * (ord(letter)-ord("A") + (not first))
        i+= 1
        first = False
    return index

class blankable(str):
    """
    A signalling class for a possible blank string
    This is because pandas will fail to read a cell if it is blank but some blank cells still need to be read
    """
    pass

class EmptyCell(Exception):
    """EmptyCell is an exception that is raised when a cell lacks a value"""
    pass

class interface:
    """
    A tool to read and write to xlsx files to garuntee datatypes, contains a pandas DataFrame
    Use interface.frame to interact with the pandas DataFrame contained within the interface
    """

    def __init__(self, frame:pandas.DataFrame):
        if not isinstance(frame, pandas.DataFrame):
            raise TypeError("Interfacer must be passed a pandas.DataFrame")
        self.frame = frame
    
    def isEmpty(self, x:int, y:int):
        """isEmpty will return false if a value is present in the cell"""
        try:
            # if the value isnan, there is no value present
            if (numpy.isnan(self.frame.at[y,x])): return True
        except KeyError:
            # if a key error occurs, there is no value in that cell
            return True
        except TypeError:
            # if a type error occurs, it is because isnan threw an error because the value was not nan, meaning a value exists
            return False
        return False

    def read(self, readtype:object, x:int, y:int):
        """
        Read will attempt to read a value at x,y and return said type
        """

        if readtype == blankable:
            if self.isEmpty(x,y): return ""
            return str(self.frame.at[y,x])
        
        if self.isEmpty(x,y): raise EmptyCell("Expected cell at y,x: "+str(y)+","+str(x))

        rawvalue = self.frame.at[y,x]

        try:
            if readtype == str:
                return str(rawvalue)
            elif readtype == int:
                return int(rawvalue)
            elif readtype == float:
                return float(rawvalue)
            elif readtype == bool:
                if type(rawvalue) == bool:
                    return rawvalue
                elif type(rawvalue) == str:
                    return rawvalue.upper()=="TRUE"
                elif type(rawvalue) in [int,float]:
                    return rawvalue>0
                elif type(rawvalue) == numpy.float64:
                    # this must be seaparate from the other numbers because the comparison will create a
                    # numpy.bool_ which will crash the json dumps
                    return bool(rawvalue>0)
        except ValueError:
            raise ValueError("Interface failed to read an item of type \""+readtype.__name__+"\" at y,x "+str(y)+","+str(x)+" due to an incorrect read type")

        raise Exception("Interface failed to read an item of type \""+readtype.__name__+"\" at y,x "+str(y)+","+str(x))

    def write(self, value:typing.Union[bool,int,float,str], x:int, y:int):
        self.frame.loc[y,x] = value

    def readIntoDict(self, readtype:object, x:int, y:int, dictionary:dict, key:str):
        """
        ReadIntoDict will attempt to read a value at x,y
        If such a value exists, it will set dict[key] to the value
        If no such value exists, it will return None
        """

        try:
            v = self.read(readtype, x, y)
            dictionary[key] = v
            return v
        except EmptyCell:
            return None

    def writeFromDict(self, x:int, y:int, dictionary:dict, key:str):
        """
        ReadIntoDict will attempt to write a value at x,y
        The value to be written is from dict[key], if no such value for the key exists, the function does nothing
        returns True when a value is written, returns False when nothing is written
        """

        try:
            v = dictionary[key]
            self.write(v, x, y)
            return True
        except KeyError:
            return False

    def save(self, filename:str, sheet:str):
        """
        Save a dataframe onto an existing sheet that has pre-existing formatting
        """
        workbook = openpyxl.load_workbook(filename=filename)
        try:
            sheet = workbook[sheet]
        except KeyError:
            sheet = workbook.create_sheet(sheet)

        # iterate over all cells and copy the data
        for y in self.frame.axes[0]:
            for x in self.frame.axes[1]:
                v = self.frame.at[y,x]
                try:
                    # NaNs should not be written
                    if numpy.isnan(v): continue
                except: pass
                # add 1 to row and column since openpyxl starts counting at 1
                sheet.cell(row=y+1, column=x+1).value = v

        workbook.save(filename = filename)

if __name__ == "__main__":
    """Test code to read/write to 12 cells of a test sheet and put the values into a dict and pull values from a dict"""
    i = interface(pandas.read_excel("./test/testinterface.xlsx", "Sheet1", header=None))
    print(i.read(str, 0, 0))
    print(i.read(int, 0, 1))
    print(i.read(float, 0, 2))
    print(i.read(bool, 0, 3))
    # print(i.read(int, 4, 0))
    a={}
    i.readIntoDict(str, 0, 0, a, "String1")
    i.readIntoDict(int, 0, 1, a, "Int1")
    i.readIntoDict(float, 0, 2, a, "Float1")
    i.readIntoDict(int, 0, 4, a, "Int2") # should not end up in the dict
    print(a)

    i.write(0, 1, 0)
    i.write("c", 1, 1)
    i.write(True, 1, 2)
    i.write(5.6, 1, 3)

    i.writeFromDict(2, 0, a, "String1")
    i.writeFromDict(2, 1, a, "Int1")
    i.writeFromDict(2, 2, a, "Float1")
    i.writeFromDict(2, 3, a, "Int2") # should not be written as Int2 should not end up in the dict

    i.frame.to_excel("./test/testinterface.xlsx", sheet_name="Sheet1", index=False, header=None)
    input("Done.")
